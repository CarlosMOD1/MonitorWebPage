#!/usr/bin/env python3
"""
reports.py — Generacion de reportes Excel con historial de yield.

Genera un workbook con 5 hojas:
  1. Summary       — metricas globales del periodo seleccionado
  2. Daily Yield   — historial diario (ultimos 30 dias)
  3. Weekly Yield  — historial semanal (ultimas 12 semanas)
  4. Monthly Yield — historial mensual (ultimos 12 meses)
  5. Top Failures  — top 50 modos de falla con distribucion
"""

from io import BytesIO
from datetime import datetime
from functools import reduce
from operator import mul
import re

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart
from openpyxl.chart.reference import Reference
from openpyxl.utils import get_column_letter


# Caracteres ilegales en XML / Excel (según spec OOXML)
_ILLEGAL_CHARS_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x84\x86-\x9f\ufffe\uffff]'
)


def _safe(val):
    """Elimina caracteres ilegales de strings antes de escribirlos en una celda."""
    if isinstance(val, str):
        return _ILLEGAL_CHARS_RE.sub('', val)
    return val


# ─────────────────────────────────────────────────────────────
# PALETA DE COLORES (tema oscuro coherente con la UI)
# ─────────────────────────────────────────────────────────────
_C = {
    "bg_dark":    "0F1923",
    "bg_header":  "0A1420",
    "bg_row_a":   "1A2B3C",
    "bg_row_b":   "142131",
    "title_blue": "2E9CDB",
    "green":      "00D4AA",
    "yellow":     "F1C40F",
    "red":        "E74C3C",
    "gray":       "8B9BAA",
    "white":      "F4F6F8",
    "border":     "2C3E50",
}


# ─────────────────────────────────────────────────────────────
# HELPERS DE ESTILO
# ─────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=None, size=10):
    return Font(bold=bold, color=(color or _C["white"]), size=size, name="Segoe UI")


def _border():
    s = Side(style="thin", color=_C["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)


def _yield_color(value):
    """Semaforo: verde >=95, amarillo >=85, rojo <85."""
    if isinstance(value, (int, float)):
        if value >= 95:
            return _C["green"]
        if value >= 85:
            return _C["yellow"]
    return _C["red"]


# ─────────────────────────────────────────────────────────────
# ESTRUCTURA DE HOJAS
# ─────────────────────────────────────────────────────────────

def _write_sheet_title(ws, title, subtitle=""):
    """Escribe fila de titulo y subtitulo con estilo en la hoja."""
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = title
    c.fill      = _fill(_C["bg_dark"])
    c.font      = _font(bold=True, size=14, color=_C["title_blue"])
    c.alignment = _left()

    if subtitle:
        ws.row_dimensions[2].height = 16
        ws.merge_cells("A2:H2")
        c2 = ws["A2"]
        c2.value     = subtitle
        c2.fill      = _fill(_C["bg_dark"])
        c2.font      = _font(size=9, color=_C["gray"])
        c2.alignment = _left()


def _write_header_row(ws, row_num, headers, col_widths=None):
    """Escribe fila de encabezados con fondo oscuro y texto azul."""
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=text)
        cell.fill      = _fill(_C["bg_header"])
        cell.font      = _font(bold=True, color=_C["title_blue"])
        cell.border    = _border()
        cell.alignment = _center()

    if col_widths:
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width


def _write_data_row(ws, row_num, values, yield_col=None):
    """Escribe una fila de datos con fondo alterno y semaforo en la columna yield."""
    bg = _C["bg_row_a"] if row_num % 2 == 0 else _C["bg_row_b"]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=_safe(val))
        cell.fill      = _fill(bg)
        cell.border    = _border()
        cell.alignment = _center()
        # Semaforo de color solo en la columna de yield
        if yield_col and col == yield_col:
            cell.font = _font(bold=True, color=_yield_color(val))
        else:
            cell.font = _font(color="D0D8E0")


# ─────────────────────────────────────────────────────────────
# SERIES DE TIEMPO
# ─────────────────────────────────────────────────────────────

def _build_time_series(df):
    """
    Calcula series de tiempo agrupadas por dia, semana y mes.
    Retorna (daily_df, weekly_df, monthly_df).
    """
    df = df.copy()
    # Asegurar que `endtime` sea datetime; si no, intentar parsear y descartar no-parseables
    if not pd.api.types.is_datetime64_any_dtype(df["endtime"]):
        df["endtime"] = pd.to_datetime(df["endtime"], errors="coerce")
    df = df.dropna(subset=["endtime"])  # eliminar filas sin fecha valida

    df["period_d"] = df["endtime"].dt.date
    df["period_w"] = df["endtime"].dt.to_period("W")
    df["period_m"] = df["endtime"].dt.to_period("M")

    def _agg(grp_col):
        agg = (
            df.groupby(grp_col)
            .agg(
                total  =("resultado", "count"),
                passed =("resultado", lambda x: (x == "PASSED").sum()),
                failed =("resultado", lambda x: (x == "FAILED").sum()),
                fpy_cnt=("failureCode", lambda x: x.str.contains(r".*-PASS-.*", na=False).sum()),
            )
            .reset_index()
        )
        agg["yield_pct"] = (agg["passed"] / agg["total"] * 100).round(1).where(agg["total"] > 0, 0)
        agg["fpy_pct"]   = (agg["fpy_cnt"] / agg["total"] * 100).round(1).where(agg["total"] > 0, 0)
        return agg

    daily   = _agg("period_d").tail(30)
    weekly  = _agg("period_w").tail(12)
    monthly = _agg("period_m").tail(12)
    return daily, weekly, monthly


# ─────────────────────────────────────────────────────────────
# HOJA 1: SUMMARY
# ─────────────────────────────────────────────────────────────

def _add_summary_sheet(wb, df, label, date_from_str, date_to_str):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20

    subtitle = (
        f"Period: {date_from_str}  →  {date_to_str}   |   "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    _write_sheet_title(ws, f"Manufacturing Report  —  {label}", subtitle)

    total   = len(df)
    passed  = int((df["resultado"] == "PASSED").sum())
    failed  = int((df["resultado"] == "FAILED").sum())
    yld_pct = round(passed / total * 100, 1) if total else 0.0
    fpy_cnt = int(df["failureCode"].str.contains(r".*-PASS-.*", na=False).sum())
    fpy_pct = round(fpy_cnt / total * 100, 1) if total else 0.0

    # RTY: producto de los yields por estacion (solo cuando hay > 1 estacion)
    rty_row = None
    station_groups = [(sid, g) for sid, g in df.groupby("stationid") if len(g) > 0]
    if len(station_groups) > 1:
        yields = [
            int((g["resultado"] == "PASSED").sum()) / len(g)
            for _, g in station_groups
        ]
        rty = round(reduce(mul, yields, 1.0) * 100, 1)
        rty_row = ("RTY % (Rolled Throughput Yield)", rty)

    metrics = [
        ("Total Tests",    total),
        ("Passed",         passed),
        ("Failed",         failed),
        ("Yield %",        yld_pct),
        ("FPY %",          fpy_pct),
    ]
    if rty_row:
        metrics.append(rty_row)
    metrics += [
        ("Period (days)",  (pd.to_datetime(date_to_str) - pd.to_datetime(date_from_str)).days + 1),
        ("Date From",      date_from_str),
        ("Date To",        date_to_str),
        ("Label",          label),
    ]

    start_row = 4
    for i, (metric, val) in enumerate(metrics, start=start_row):
        # Columna A: nombre de la metrica
        ca = ws.cell(i, 1, metric)
        ca.fill      = _fill(_C["bg_header"])
        ca.font      = _font(bold=True, color=_C["title_blue"])
        ca.border    = _border()
        ca.alignment = _left()

        # Columna B: valor
        cb = ws.cell(i, 2, val)
        cb.fill   = _fill(_C["bg_row_b"])
        cb.border = _border()
        cb.alignment = _center()
        if metric in ("Yield %", "FPY %", "RTY % (Rolled Throughput Yield)"):
            cb.font = _font(bold=True, color=_yield_color(val))
        else:
            cb.font = _font()


# ─────────────────────────────────────────────────────────────
# HOJAS 2-4: SERIES DE TIEMPO
# ─────────────────────────────────────────────────────────────

def _add_daily_sheet(wb, daily):
    ws = wb.create_sheet("Daily Yield")
    _write_sheet_title(ws, "Daily Yield History", "Last 30 days")
    _write_header_row(ws, 3,
        ["Date", "Total Tests", "Passed", "Failed", "Yield %", "FPY %"],
        [14, 14, 12, 12, 12, 12])
    for i, row in enumerate(daily.itertuples(), start=4):
        _write_data_row(ws, i,
            [str(row.period_d), int(row.total), int(row.passed), int(row.failed),
             float(row.yield_pct), float(row.fpy_pct)],
            yield_col=5)


def _add_weekly_sheet(wb, weekly):
    ws = wb.create_sheet("Weekly Yield")
    _write_sheet_title(ws, "Weekly Yield History", "Last 12 weeks")
    _write_header_row(ws, 3,
        ["Week", "Total Tests", "Passed", "Failed", "Yield %", "FPY %"],
        [18, 14, 12, 12, 12, 12])
    for i, row in enumerate(weekly.itertuples(), start=4):
        _write_data_row(ws, i,
            [str(row.period_w), int(row.total), int(row.passed), int(row.failed),
             float(row.yield_pct), float(row.fpy_pct)],
            yield_col=5)


def _add_monthly_sheet(wb, monthly):
    ws = wb.create_sheet("Monthly Yield")
    _write_sheet_title(ws, "Monthly Yield History", "Last 12 months")
    _write_header_row(ws, 3,
        ["Month", "Total Tests", "Passed", "Failed", "Yield %", "FPY %"],
        [14, 14, 12, 12, 12, 12])
    for i, row in enumerate(monthly.itertuples(), start=4):
        _write_data_row(ws, i,
            [str(row.period_m), int(row.total), int(row.passed), int(row.failed),
             float(row.yield_pct), float(row.fpy_pct)],
            yield_col=5)


# ─────────────────────────────────────────────────────────────
# HOJA 5: TOP FAILURES
# ─────────────────────────────────────────────────────────────

def _add_failures_sheet(wb, df):
    ws = wb.create_sheet("Top Failures")
    _write_sheet_title(ws, "Top 50 Failure Modes")
    _write_header_row(ws, 3,
        ["Failure Type", "Station", "Product", "Count", "% of Failures"],
        [36, 22, 14, 10, 16])

    df_fail     = df[df["resultado"] == "FAILED"]
    total_fails = len(df_fail)
    if df_fail.empty:
        return

    agg = (
        df_fail.groupby(["tipoFalla", "stationName", "producto"])
        .size()
        .reset_index(name="count")
        .sort_values("count", ascending=False)
        .head(50)
    )
    for i, row in enumerate(agg.itertuples(), start=4):
        pct = round(row.count / total_fails * 100, 1) if total_fails else 0.0
        _write_data_row(ws, i, [row.tipoFalla, row.stationName, row.producto, int(row.count), pct])


def _add_rty_sheet(wb, df, min_units=5):
    """Añade hoja con impacto de cada estación en el RTY.
    Calcula FPY usando solo el PRIMER intento por unidad (`prevQr`).
    """
    ws = wb.create_sheet("RTY by Station")
    _write_sheet_title(ws, "RTY by Station", f"First-attempt FPY (min_units={min_units})")
    headers = ["Station ID", "Station Name", "Product", "Units (first)", "First Pass", "Not First Pass", "FPY %", "Included in RTY"]
    _write_header_row(ws, 3, headers, [12, 28, 12, 14, 12, 14, 10, 16])

    row_num = 4
    groups = [g for _, g in df.groupby(["stationid", "stationName", "producto"]) if len(g) > 0]
    for g in groups:
        first = g.sort_values("endtime").drop_duplicates(subset="prevQr", keep="first")
        total_first = len(first)
        first_pass = int((first["resultado"] == "PASSED").sum())
        not_first = total_first - first_pass
        fpy_pct = round(first_pass / total_first * 100, 1) if total_first else 0.0
        included = total_first >= min_units

        _write_data_row(ws, row_num, [
            int(g["stationid"].iloc[0]),
            str(g["stationName"].iloc[0]),
            g["producto"].iloc[0] if "producto" in g.columns else "",
            total_first,
            first_pass,
            not_first,
            fpy_pct,
            "Yes" if included else "No",
        ], yield_col=7)
        row_num += 1


# ─────────────────────────────────────────────────────────────
# GRAFICAS DE LINEA (opcionales; se omiten si falla openpyxl)
# ─────────────────────────────────────────────────────────────

def _add_yield_chart(wb, sheet_name, chart_title, data_start_row, data_rows):
    """Inserta un grafico de linea de Yield % en la hoja indicada."""
    try:
        if data_rows < 2:
            return
        ws    = wb[sheet_name]
        chart = LineChart()
        chart.title          = chart_title
        chart.style          = 10
        chart.y_axis.title   = "Yield %"
        chart.x_axis.title   = "Period"
        chart.height         = 10
        chart.width          = 20

        # Columna 5 = Yield %, columna 1 = etiqueta de periodo
        data_ref = Reference(ws, min_col=5, min_row=data_start_row,
                             max_row=data_start_row + data_rows - 1)
        cats_ref = Reference(ws, min_col=1, min_row=data_start_row,
                             max_row=data_start_row + data_rows - 1)
        chart.add_data(data_ref)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, f"H{data_start_row}")
    except Exception:
        pass  # Los graficos son opcionales; si fallan no interrumpimos la descarga


# ─────────────────────────────────────────────────────────────
# FUNCION PRINCIPAL
# ─────────────────────────────────────────────────────────────

def generate_excel_report(df, label, date_from_str, date_to_str):
    """
    Genera un archivo Excel en memoria y retorna un BytesIO listo para Flask.

    Hojas incluidas:
      - Summary      : metricas globales + RTY si aplica
      - Daily Yield  : historial diario (30 dias) con semaforo
      - Weekly Yield : historial semanal (12 semanas) con semaforo
      - Monthly Yield: historial mensual (12 meses) con semaforo
      - Top Failures : top 50 modos de falla ordenados por frecuencia

    Parametros:
      df            — DataFrame ya filtrado (producto/estacion/fechas)
      label         — Etiqueta descriptiva (ej. 'SPSF', 'SPSF — SMT Oven')
      date_from_str — Fecha inicio en formato YYYY-MM-DD
      date_to_str   — Fecha fin   en formato YYYY-MM-DD
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # eliminar hoja vacia por defecto

    daily, weekly, monthly = _build_time_series(df)

    _add_summary_sheet(wb, df, label, date_from_str, date_to_str)
    _add_daily_sheet(wb, daily)
    _add_weekly_sheet(wb, weekly)
    _add_monthly_sheet(wb, monthly)
    _add_failures_sheet(wb, df)
    _add_rty_sheet(wb, df)

    # Graficas de tendencia
    _add_yield_chart(wb, "Daily Yield",   "Daily Yield Trend",   4, len(daily))
    _add_yield_chart(wb, "Weekly Yield",  "Weekly Yield Trend",  4, len(weekly))
    _add_yield_chart(wb, "Monthly Yield", "Monthly Yield Trend", 4, len(monthly))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
